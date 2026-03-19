import datetime as dt
import time
from io import BytesIO
from pathlib import Path
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
import xlsxwriter  # type: ignore

# ──────────────────────────────────────────────────────────────────────────────
# 0) DEFINICIÓN CORREGIDA DE parse_date (al inicio del script)
# ──────────────────────────────────────────────────────────────────────────────
import pandas as pd
from datetime import date, datetime as dt_datetime

def parse_date(val):
    """
    Convierte val (date, datetime, pandas.Timestamp o str) a date.
    """
    # Si ya es un date puro (no datetime), lo devolvemos
    if isinstance(val, date) and not isinstance(val, dt_datetime):
        return val
    # Si es un datetime.datetime de Python
    if isinstance(val, dt_datetime):
        return val.date()
    # Si es un pandas.Timestamp
    if isinstance(val, pd.Timestamp):
        return val.to_pydatetime().date()
    # Si es string, probamos formatos
    s = str(val).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return dt_datetime.strptime(s, fmt).date()
        except Exception:
            continue
    # Si no encaja, error claro
    raise ValueError(f"Formato de fecha inválido: {val!r}")

# ------------------------------------------------------------------
# ————— Autenticación con roles —————
if "logged" not in st.session_state:
    st.session_state.logged = False

if not st.session_state.logged:
    st.title("🔒 Iniciar sesión")
    user_input = st.text_input("Usuario")
    pwd_input  = st.text_input("Contraseña", type="password")
    if st.button("Entrar"):
        creds = st.secrets["credentials"]
        match = next((info for info in creds.values() if info["username"] == user_input), None)

        if match and pwd_input == match["password"]:
            st.session_state.logged = True
            st.session_state.user   = match["username"]
            st.session_state.role   = match["role"]
            # Rerun automático con la misma sesión
            st.rerun()
        else:
            st.error("Usuario o contraseña inválida")

    # Mientras no haya validado, detener todo lo demás
    st.stop()



# ------------------------------------------------------------------
# 2) Configuración global y helpers
# ------------------------------------------------------------------
BASE_DIR = Path(__file__).parent
EXCEL   = BASE_DIR / "Recursos.xlsx"
SHEET    = 'Reservas'
DATE_FMT = '%d/%m/%Y'
BURGUNDY = '#800000'
BLOQUES  = [
    (dt.time(8, 0),  dt.time(9, 30)),
    (dt.time(9, 45), dt.time(11, 15)),
    (dt.time(11, 30), dt.time(13, 0)),
    (dt.time(14, 0),  dt.time(15, 30)),
    (dt.time(15, 45), dt.time(16, 30)),
    (dt.time(16, 30), dt.time(18, 30)),
]


def as_time(val):
    if isinstance(val, dt.time):
        return val
    for fmt in ('%H:%M:%S', '%H:%M'):
        try:
            return dt.datetime.strptime(str(val), fmt).time()
        except ValueError:
            continue
    raise ValueError(f'Formato de hora inválido: {val}')

def overlap(hi1, hf1, hi2, hf2):
    return max(hi1, hi2) < min(hf1, hf2)

from pathlib import Path
import pandas as pd

# ------------------------------------------------------------------
# 3) Recarga de datos inicial y listas dinámicas
# ------------------------------------------------------------------
from pathlib import Path
import zipfile
from openpyxl import Workbook, load_workbook

# Nombre de tu fichero y de la hoja
EXCEL = "Recursos.xlsx"
SHEET = "Reservas"

# Las columnas esperadas en tu base de datos
DB_COLS = [
    "Fecha", "Hora inicio", "Hora fin",
    "Profesor", "Curso", "Recurso", "Observaciones"
]

# Asegúrate de que EXCEL termina en .xlsx
if not EXCEL.lower().endswith('.xlsx'):
    raise ValueError(f"El fichero de base de datos debe ser .xlsx, no {EXCEL!r}")

def init_empty_db(path: str, sheet_name: str):
    """
    Crea el .xlsx y escribe en la hoja sheet_name
    la primera fila con los encabezados DB_COLS.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    # Escribir encabezados
    for col_idx, header in enumerate(DB_COLS, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    wb.save(path)

# Si no existe el fichero, créalo con los encabezados
if not Path(EXCEL).exists():
    init_empty_db(EXCEL, SHEET)
else:
    # Si abre mal o falta la hoja, reemplázalo
    try:
        with zipfile.ZipFile(EXCEL, 'r'):
            pass
        wb = load_workbook(EXCEL)
        if SHEET not in wb.sheetnames:
            init_empty_db(EXCEL, SHEET)
    except (zipfile.BadZipFile, KeyError):
        init_empty_db(EXCEL, SHEET)

# Carga segura de datos
df = pd.read_excel(EXCEL, sheet_name=SHEET, dtype=str).fillna('')

# Listas dinámicas
def recalc_lists(df: pd.DataFrame) -> tuple[list,str,list[str]]:
    """
    Recalcula las listas de Profesores, Cursos y Recursos
    a partir de las hojas correspondientes en el archivo EXCEL
    y de lo que ya hay en df.
    """
    # Cargamos el libro y sus nombres de hoja
    xls = pd.ExcelFile(EXCEL)
    sheets = xls.sheet_names

    # — Profesores —
    if 'Profesores' in sheets:
        prof_df = pd.read_excel(EXCEL, sheet_name='Profesores').fillna('')
        profs = prof_df.iloc[:, 0].astype(str).tolist()
    else:
        profs = []
    # Añadimos los que ya existen en df
    profs = sorted(set(profs) | set(df['Profesor'].dropna().astype(str).unique()))

    # — Cursos —
    if 'Cursos' in sheets:
        cursos_df = pd.read_excel(EXCEL, sheet_name='Cursos').fillna('')
        cursos = cursos_df.iloc[:, 0].astype(str).tolist()
    else:
        cursos = []
    cursos = sorted(set(cursos) | set(df['Curso'].dropna().astype(str).unique()))

    # — Recursos —
    if 'Recursos' in sheets:
        recs_df = pd.read_excel(EXCEL, sheet_name='Recursos').fillna('')
        recs = recs_df.iloc[:, 0].astype(str).tolist()
    else:
        recs = []
    # También incluimos los recursos ya usados en df (cadena "a, b, c")
    usados = []
    for cell in df['Recurso'].dropna().astype(str):
        usados.extend([r.strip() for r in cell.split(',')])
    recursos = sorted(set(recs) | set(usados))

    return profs, cursos, recursos


# ------------------------------------------------------------------
# 4) Funciones de guardado e informes
# ------------------------------------------------------------------
def atomic_save(df_save: pd.DataFrame) -> None:
    for _ in range(5):
        try:
            wb = load_workbook(EXCEL)
            if SHEET in wb.sheetnames:
                wb.remove(wb[SHEET])
            ws = wb.create_sheet(SHEET, 0)
            for c, col in enumerate(df_save.columns, 1):
                ws.cell(1, c, col)
            for r, row in enumerate(df_save.itertuples(index=False), 2):
                for c, val in enumerate(row, 1):
                    ws.cell(r, c, val)
            wb.save(EXCEL)
            break
        except PermissionError:
            time.sleep(0.2)
    else:
        toast(
            "❌ No se pudo guardar el archivo. "
            "Cierra cualquier otra aplicación que lo tenga abierto.",
            "error"
        )

def build_report(df_report: pd.DataFrame) -> bytes:
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="xlsxwriter") as writer:
        # 1) Hoja de reservas
        df_report.to_excel(writer, sheet_name="Reservas", index=False)

        # 2) Hoja Uso recurso
        recurso = df_report["Recurso"].str.split(", ").explode()
        uso_rec = recurso.value_counts().rename("Reservas")
        uso_rec.to_excel(writer, sheet_name="Uso recurso")

        # 3) Hoja Uso profesor
        uso_prof = df_report["Profesor"].value_counts().rename("Reservas")
        uso_prof.to_excel(writer, sheet_name="Uso profesor")

        # 4) Hoja de gráficos
        workbook = writer.book
        chart_ws = workbook.add_worksheet("Gráficos")

        # Pie chart: Uso recurso
        chart1 = workbook.add_chart({"type": "pie"})
        n_rec   = uso_rec.size
        chart1.add_series({
            "name":       "Uso recurso",
            "categories": f"='Uso recurso'!$A$2:$A${n_rec+1}",
            "values":     f"='Uso recurso'!$B$2:$B${n_rec+1}",
            "data_labels": {"percentage": True},
        })
        chart1.set_title({"name": "Uso recurso"})
        chart_ws.insert_chart("A1", chart1)

        # Pie chart: Uso profesor
        chart2 = workbook.add_chart({"type": "pie"})
        n_prof  = uso_prof.size
        chart2.add_series({
            "name":       "Uso profesor",
            "categories": f"='Uso profesor'!$A$2:$A${n_prof+1}",
            "values":     f"='Uso profesor'!$B$2:$B${n_prof+1}",
            "data_labels": {"percentage": True},
        })
        chart2.set_title({"name": "Uso profesor"})
        chart_ws.insert_chart("H1", chart2)

    return buf.getvalue()

def build_ics(df_events: pd.DataFrame, prof: str) -> str:
    now = dt.datetime.utcnow()
    lines = [
        'BEGIN:VCALENDAR', 'VERSION:2.0',
        'PRODID:-//CAV//Horario Enlaces//ES', 'CALSCALE:GREGORIAN'
    ]
    for idx, row in df_events.iterrows():
        d = parse_date(row['Fecha'])
        hi = as_time(row['Hora inicio']); hf = as_time(row['Hora fin'])
        sd = dt.datetime.combine(d, hi); ed = dt.datetime.combine(d, hf)
        uid = f"{prof}-{idx}@cav.cl"
        lines += [
            'BEGIN:VEVENT',
            f'UID:{uid}',
            f'DTSTAMP:{now.strftime("%Y%m%dT%H%M%SZ")}',
            f'DTSTART;TZID=America/Santiago:{sd.strftime("%Y%m%dT%H%M%S")}',
            f'DTEND;TZID=America/Santiago:{ed.strftime("%Y%m%dT%H%M%S")}',
            f"SUMMARY:Reserva {row['Recurso']} ({row['Curso']})",
            f"DESCRIPTION:{row.get('Observaciones', '')}",
            'END:VEVENT'
        ]
    lines.append('END:VCALENDAR')
    return '\n'.join(lines)

# ------------------------------------------------------------------
# 5) Configuración de página y estilos
# ------------------------------------------------------------------
st.set_page_config('📅 HORARIO ENLACES CAV 💻', page_icon='', layout='wide')
CSS = f"""
html, body, [data-testid='stApp'] {{ background:#fff; color:#000; }}
.stButton>button, .stDownloadButton>button {{ background:{BURGUNDY}; color:#fff; }}
/* section[data-testid='stSidebar'] {{ display:none; }} */
[data-testid='stAppViewContainer']>div {{ padding:0 3rem; }}
label {{ color:#000!important; font-weight:600; }}
.stTableContainer .stTable th, .stTableContainer .stTable td {{
    border:1px solid #000!important; color:#000!important;
}}
[data-testid='stDataEditorContainer'] * {{ color:#000!important; }}
div[role='tablist'] > button[role='tab'] {{ color:#000!important; }}

@media (prefers-color-scheme: dark) {{
  html, body, [data-testid='stApp'] {{ background:#000; color:#fff; }}
  .stButton>button, .stDownloadButton>button {{ background:#5c0000; color:#fff; }}
  label {{ color:#fff!important; }}
  .stTableContainer .stTable th, .stTableContainer .stTable td {{
      border:1px solid #fff!important; color:#fff!important;
  }}
  [data-testid='stDataEditorContainer'] * {{ color:#fff!important; }}
  div[role='tablist'] > button[role='tab'] {{ color:#fff!important; }}
}}
"""

st.markdown(f"<style>{CSS}</style>", unsafe_allow_html=True)
#st.markdown(
#    f"<h1 style='text-align:center;font-size:2rem;color:{BURGUNDY};'>📅 HORARIO ENLACES CAV 💻</h1>",
#    unsafe_allow_html=True
#)

# ------------------------------------------------------------------
# 6) Función toast
# ------------------------------------------------------------------
def toast(msg: str, kind: str = 'info') -> None:
    icons = {
        'success': '✅',
        'error':   '❌',
        'warning': '⚠️',
        'info':    'ℹ️',
    }
    try:
        st.toast(msg, icon=icons.get(kind))
    except Exception:
        getattr(st, kind)(msg)

# ------------------------------------------------------------------
import streamlit.components.v1 as components
st.markdown(
    """
    <style>
      /* 1) Hacer que la Sidebar quede fija al hacer scroll */
      section[data-testid="stSidebar"] {
        position: sticky;
        top: 0;
        height: 100vh;
        overflow-y: auto;
      }

      /* 2) Ajustar ancho de la Sidebar si lo necesitas */
      section[data-testid="stSidebar"] {
        width: 260px !important;
      }

      /* 3) Aumentar el tamaño de letra del menú de navegación (radio buttons) */
      section[data-testid="stSidebar"] .stRadio label {
        font-size: 2rem !important;
        font-weight: 500 !important;
      }
    </style>
    """,
    unsafe_allow_html=True
)
# 3) Logo con ancho fijo (ajusta el valor a tu gusto)
st.sidebar.image("logo_CAV_2021-1.png", width=200)  # <- aquí cambias el tamaño
# 1) Datos de sesión
st.sidebar.markdown(f"**✅ Usuario:** {st.session_state.user}")
# 4) Título con tamaño de fuente personalizado
st.sidebar.markdown(
    "<h3 style='color:#800000; text-align:center; font-size:1.2rem; margin:0.5rem 0;'>"
    "💻HORARIO ENLACES CAV"
    "</h3>",
    unsafe_allow_html=True
)
# 5) Menú de navegación
role = st.session_state.role
if role == 'admin':
    pages = ['▶ Registrar','📂 Base datos','📅 Semana','🔧 Mantenimiento']
elif role == 'profesor':
    pages = ['▶ Registrar','📅 Semana']
else:
    pages = ['▶ Registrar']
page = st.sidebar.radio("📂 Navegar a:", pages, index=0)
# ——— Sidebar fijo con usuario, logout, logo y navegación ———



# 2) Botón Cerrar sesión
if st.sidebar.button("🚪  Cerrar sesión", use_container_width=True):
    for k in ["logged","user","role"]:
        st.session_state.pop(k, None)
    st.rerun()
# ————————————————————————————————————————————————

# ------------------------------------------------------------------
# 8) Sección ▶ Registrar (solo si page == '▶ Registrar')
# ------------------------------------------------------------------
if page == '▶ Registrar':
    st.markdown(
        "<h2 style='color:#000;font-size:1.5rem;text-align:center'>⏩ Registrar nueva reserva</h2>",
        unsafe_allow_html=True
    )

    # — Carrusel + KPI — solo aquí —
    col_car, col_kpi = st.columns([2, 1], gap="small")

    with col_car:
        hoy   = dt.date.today()
        lunes = hoy - dt.timedelta(days=hoy.weekday())
        dias  = [lunes + dt.timedelta(days=i) for i in range(5)]
        nombres = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes"]
        dur     = 4

        slides = []
        for i, d in enumerate(dias):
            title = f"{nombres[i]} {d.strftime('%d/%m')}"
            sel   = df[df["Fecha"].apply(parse_date) == d]
            body  = (
                sel[["Hora inicio", "Hora fin", "Profesor", "Curso", "Recurso"]]
                .to_html(index=False)
                if not sel.empty else
                "<p>Sin reservas</p>"
            )
            slides.append(f"<div class='slide'><h5>{title}</h5>{body}</div>")

        html = f"""
        <style>
          .carousel {{ position: relative; width:100%; height:160px; overflow:hidden;
                        border:1px solid #ddd; border-radius:8px; }}
          .slide    {{ position:absolute; width:100%; height:100%; top:0; left:0;
                        opacity:0; animation: carousel {len(dias)*dur}s infinite; }}
          {"".join(f".slide:nth-child({i+1}){{animation-delay:{i*dur}s}} " for i in range(len(dias)))}
          @keyframes carousel {{0%{{opacity:0}} 5%{{opacity:1}} 20%{{opacity:1}}
                                25%{{opacity:0}} 100%{{opacity:0}}}}
          .slide h5    {{ margin:0.2rem; color:{BURGUNDY}; font-size:0.8rem;
                         text-align:center; }}
          .slide table {{ margin:0.2rem auto; width:95%; border-collapse:collapse;
                         font-size:0.6rem; text-align:center; }}
          .slide th,
          .slide td    {{ border:0.5px solid #000; padding:2px; }}
        </style>
        <div class="carousel">{''.join(slides)}</div>
        """
        components.html(html, height=160)

    with col_kpi:
        hoy       = dt.date.today()
        total     = len(df)
        hoy_count = (df["Fecha"].apply(parse_date) == hoy).sum()

        # filtramos reservas de hoy o futuras
        future = df[df["Fecha"].apply(parse_date) >= hoy].copy()
        if not future.empty:
            # parseamos fecha y hora a objetos
            future["Fecha_dt"] = future["Fecha"].apply(parse_date)
            future["Hora_dt"]  = future["Hora inicio"].apply(as_time)
            # ordenamos por fecha+hora
            future = future.sort_values(["Fecha_dt", "Hora_dt"])
            fila = future.iloc[0]
            # formateamos correctamente
            fecha_str = fila["Fecha_dt"].strftime(DATE_FMT)
            hora_str  = fila["Hora_dt"].strftime("%H:%M")
            valor = f"{fecha_str} {hora_str}"
        else:
            valor = None

        st.markdown(
            f"<p style='font-size:1.3rem;margin:0'>{total}</p>"
            "<p style='font-size:0.8rem;color:gray;margin:0;'>🗓️ Total reservas</p>",
            unsafe_allow_html=True
        )
        st.markdown(
            f"<p style='font-size:1.1rem;margin:0'>{hoy_count}</p>"
            "<p style='font-size:0.8rem;color:gray;margin:0;'>📅 Reservas hoy</p>",
            unsafe_allow_html=True
        )
        st.markdown(
            f"<p style='font-size:1.1rem;margin:0'>{valor if valor else '–'}</p>"
            "<p style='font-size:0.8rem;color:gray;margin:0;'>⏰ Próxima reserva</p>",
            unsafe_allow_html=True
        )

    # — Formulario registrar —  
    PROFESORES, CURSOS, RECURSOS = recalc_lists(df)

    def course_key(c: str) -> tuple[int, str]:
        cu = c.upper()
        if 'BÁSIC' in cu: return (0, cu)
        if 'MEDIO' in cu: return (1, cu)
        if 'DIF'   in cu: return (2, cu)
        return (3, cu)

    CURSOS = sorted(CURSOS, key=course_key)

    with st.form('reserva_form'):
        c1, c2 = st.columns(2, gap="small")
        with c1:
            fecha      = st.date_input('Fecha inicial', dt.date.today(), format='DD/MM/YYYY')
            recurrente = st.checkbox("🔁 Hacer esta reserva recurrente")
            if recurrente:
                freq = st.selectbox("Frecuencia", ["Semanal", "Diaria"])
                if freq == "Semanal":
                    dias_sem = st.multiselect(
                        "Días de la semana",
                        nombres,
                        default=[nombres[fecha.weekday()]]
                    )
                fecha_fin = st.date_input(
                    "Repetir hasta",
                    fecha + dt.timedelta(weeks=4),
                    help="Fecha límite de la recurrencia"
                )
            hi = st.time_input('Hora inicio', BLOQUES[0][0])
            hf = st.time_input('Hora fin',    BLOQUES[0][1])

        with c2:
            prof   = st.selectbox('Profesor', PROFESORES)
            curso  = st.selectbox('Curso',    CURSOS)

            # comprobación de mantenimientos
            unavail = []
            df_mant = pd.read_excel(EXCEL, sheet_name='Mantenimientos').fillna('')
            if not df_mant.empty:
                c0  = next((col for col in df_mant.columns if 'inicio' in col.lower()), None)
                c1_ = next((col for col in df_mant.columns if 'fin'    in col.lower()), None)
                if c0 and c1_:
                    df_mant['Inicio'] = df_mant[c0].apply(parse_date)
                    df_mant['Fin']    = df_mant[c1_].apply(parse_date)
                    unavail = df_mant[
                        (df_mant['Inicio'] <= fecha) & (df_mant['Fin'] >= fecha)
                    ]['Recurso'].astype(str).tolist()
            if unavail:
                st.warning(f"⚠️ Recursos en mantenimiento: {', '.join(unavail)}")
            available = [r for r in RECURSOS if r not in unavail]
            recs      = st.multiselect('Recursos', available)

        obs       = st.text_area('Observaciones', height=68)
        submitted = st.form_submit_button('💾 Guardar reserva')

    if submitted:
        # — Validaciones —
        if hi >= hf:
            toast('Hora inicio debe ser antes de fin.', 'warning')
            st.stop()
        if not recs:
            toast('Selecciona al menos un recurso.', 'warning')
            st.stop()

        # — Generar fechas según recurrencia —
        fechas = []
        if recurrente:
            if freq == "Diaria":
                d = fecha
                while d <= fecha_fin:
                    fechas.append(d)
                    d += dt.timedelta(days=1)
            else:  # Semanal
                weekday_map = {n: i for i, n in enumerate(nombres)}
                sel_nums    = [weekday_map[d] for d in dias_sem]
                d = fecha
                while d <= fecha_fin:
                    if d.weekday() in sel_nums:
                        fechas.append(d)
                    d += dt.timedelta(days=1)
        else:
            fechas = [fecha]

        # — Construir DataFrame de nuevas reservas —
        rows = []
        for d in fechas:
            rows.append({
                "Fecha":         d.strftime(DATE_FMT),
                "Hora inicio":   hi.strftime("%H:%M"),
                "Hora fin":      hf.strftime("%H:%M"),
                "Profesor":      prof,
                "Curso":         curso,
                "Recurso":       ", ".join(recs),
                "Observaciones": obs.strip(),
            })
        new = pd.DataFrame(rows)

        # — Guardar y notificar —
        df = pd.concat([df, new], ignore_index=True)
        atomic_save(df)
        toast(f"✅ {len(new)} reservas creadas", "success")
        # st.experimental_rerun()  <-- comentado si tu versión no lo soporta
 
# ------------------------------------------------------------------
# 9) Sección 📂 Base datos (solo admin)
# ------------------------------------------------------------------
elif page == '📂 Base datos' and role == 'admin':
    st.markdown(
        "<h2 style='color:#000;font-size:1.5rem; text-align:center'>📂 Base datos de reservas</h2>",
        unsafe_allow_html=True
    )

    # 1) Recalcular opciones de select
    PROFESORES, CURSOS, RECURSOS = recalc_lists(df)

    # 2) Filtros dinámicos en un expander
    with st.expander("🔍 Filtros", expanded=True):
        c1, c2, c3, c4 = st.columns(4)
        min_fecha = df["Fecha"].apply(parse_date).min()
        max_fecha = df["Fecha"].apply(parse_date).max()
        with c1:
            fecha_min, fecha_max = st.date_input(
                "Rango fechas",
                value=[min_fecha, max_fecha],
                format="DD/MM/YYYY"
            )
        with c2:
            profesores_sel = st.multiselect(
                "Profesores",
                options=PROFESORES,
                default=[],
                placeholder="Todos"
            )
        with c3:
            cursos_sel = st.multiselect(
                "Cursos",
                options=CURSOS,
                default=[],
                placeholder="Todos"
            )
        with c4:
            recursos_sel = st.multiselect(
                "Recursos",
                options=RECURSOS,
                default=[],
                placeholder="Todos"
            )

    # 3) Aplicar filtros (si la lista está vacía, ignoro ese criterio)
    df_work = df.copy()
    df_work["Fecha_dt"] = df_work["Fecha"].apply(parse_date)
    mask_fecha = (
        (df_work["Fecha_dt"] >= fecha_min) &
        (df_work["Fecha_dt"] <= fecha_max)
    )
    mask_prof = df_work["Profesor"].isin(profesores_sel) if profesores_sel else True
    mask_cur  = df_work["Curso"].isin(cursos_sel)         if cursos_sel    else True
    mask_rec  = df_work["Recurso"].isin(recursos_sel)     if recursos_sel  else True

    df_filt = df_work[mask_fecha & mask_prof & mask_cur & mask_rec].drop(columns="Fecha_dt")
    total, filtrados = len(df), len(df_filt)
    st.markdown(f"Mostrando **{filtrados}** de **{total}** registros")
    if filtrados == 0:
        st.warning("No hay registros que cumplan esos criterios.")

    # 4) Preparar editor: parseo de fechas/horas
    editor = df_filt.copy()
    editor['Fecha']       = editor['Fecha'].apply(parse_date)
    editor['Hora inicio'] = editor['Hora inicio'].apply(as_time)
    editor['Hora fin']    = editor['Hora fin'].apply(as_time)

    # 5) Column config con pickers y selectboxes
    cfg = {
        'Fecha':         st.column_config.DateColumn('Fecha', format='DD/MM/YYYY'),
        'Hora inicio':   st.column_config.TimeColumn('Hora inicio', format='HH:mm'),
        'Hora fin':      st.column_config.TimeColumn('Hora fin', format='HH:mm'),
        'Profesor':      st.column_config.SelectboxColumn('Profesor', options=PROFESORES),
        'Curso':         st.column_config.SelectboxColumn('Curso',    options=CURSOS),
        'Recurso':       st.column_config.SelectboxColumn('Recurso',   options=RECURSOS),
        'Observaciones': st.column_config.TextColumn('Observaciones'),
    }

    edited = st.data_editor(
        editor,
        hide_index=False,               # mantenemos índice original para merge
        use_container_width=True,
        column_config=cfg,
        height=400
    )

    st.markdown("---")

    # 6) Detección de solapamientos por recurso/horario
    def detect_conflicts(df_check: pd.DataFrame) -> list[tuple]:
        conflicts = []
        for res in RECURSOS:
            grp = df_check[df_check['Recurso'] == res]
            for fecha, g in grp.groupby('Fecha'):
                slots = sorted(
                    [(r['Hora inicio'], r['Hora fin'], idx) for idx, r in g.iterrows()],
                    key=lambda x: x[0]
                )
                for (h1, h2, i1), (h3, h4, i2) in zip(slots, slots[1:]):
                    if overlap(h1, h2, h3, h4):
                        conflicts.append((res, fecha, i1, i2))
        return conflicts

    # 7) Detección de conflictos con mantenimientos
    mant_df = pd.read_excel(EXCEL, sheet_name='Mantenimientos').fillna('')
    if not mant_df.empty:
        mant_df['FechaInicio_dt'] = mant_df['FechaInicio'].apply(parse_date)
        mant_df['FechaFin_dt']    = mant_df['FechaFin'].apply(parse_date)
        mant_df['HoraInicio_t']   = mant_df['HoraInicio'].apply(as_time)
        mant_df['HoraFin_t']      = mant_df['HoraFin'].apply(as_time)

    def detect_maintenance_conflicts(df_check: pd.DataFrame) -> list[tuple]:
        m_conflicts = []
        for idx, r in df_check.iterrows():
            res = r['Recurso']
            fecha = r['Fecha']
            hi_r, hf_r = r['Hora inicio'], r['Hora fin']
            for _, m in mant_df[mant_df['Recurso'] == res].iterrows():
                if m['FechaInicio_dt'] <= fecha <= m['FechaFin_dt']:
                    if overlap(hi_r, hf_r, m['HoraInicio_t'], m['HoraFin_t']):
                        m_conflicts.append((idx, res, fecha))
        return m_conflicts

    # 8) Botones Guardar / Eliminar con validación
    c1, c2 = st.columns(2, gap="large")
    with c1:
        if st.button('💾 Guardar cambios', key='save_db_edits', use_container_width=True):
            # Merge de cambios en df original
            df_updated = df.copy()
            for idx, row in edited.iterrows():
                df_updated.loc[idx] = row

            confs      = detect_conflicts(df_updated)
            mast_confs = detect_maintenance_conflicts(df_updated)

            if confs or mast_confs:
                for idx, res, fecha in mast_confs:
                    toast(
                        f"⚠️ El recurso {res} está en mantenimiento el {fecha.strftime(DATE_FMT)}.",
                        "error"
                    )
                for res, fecha, i1, i2 in confs:
                    toast(
                        f"⚠️ Conflicto en {res} el {fecha.strftime(DATE_FMT)} "
                        f"entre registros {i1} y {i2}.",
                        "error"
                    )
                st.error("Corrige los conflictos antes de guardar.")
            else:
                atomic_save(df_updated)
                toast('✅ Cambios guardados.', 'success')
                try:
                    st.experimental_rerun()
                except AttributeError:
                    import time
                    st.query_params = {"_refresh": int(time.time())}

    with c2:
        to_drop = st.multiselect(
            'Seleccionar registros a eliminar',
            options=edited.index,
            key='drop_db'
        )
        if to_drop and st.button('🗑️ Eliminar registros', key='delete_db', use_container_width=True):
            df_new     = df.copy().drop(index=to_drop).reset_index(drop=True)
            confs      = detect_conflicts(df_new)
            mast_confs = detect_maintenance_conflicts(df_new)
            if confs or mast_confs:
                st.error("Aún hay conflictos tras eliminar.")
            else:
                atomic_save(df_new)
                toast('✅ Registros eliminados.', 'success')
                try:
                    st.experimental_rerun()
                except AttributeError:
                    import time
                    st.query_params = {"_refresh": int(time.time())}

    st.markdown("---")

    # 9) Informe Excel + Calendario
    col3, col4 = st.columns(2, gap="large")
    with col3:
        rpt = build_report(df)
        st.download_button(
            label='📥 Informe Excel',
            data=rpt,
            file_name='informe_reservas.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            use_container_width=True
        )
    with col4:
        prof_cal = st.selectbox('Seleccionar profesor (.ics)', PROFESORES, key='export_cal')
        ics_data = build_ics(df[df['Profesor'] == prof_cal], prof_cal)
        st.download_button(
            label='📅 Descargar calendario',
            data=ics_data,
            file_name=f'{prof_cal}_reservas.ics',
            mime='text/calendar',
            use_container_width=True
        )

# ------------------------------------------------------------------
# 10) Sección 📅 Semana (admin y profesor)
# ------------------------------------------------------------------
elif page == '📅 Semana':
    st.markdown(
        "<h2 style='color:#000;font-size:1.5rem;text-align:center'>📅 Vista semanal</h2>",
        unsafe_allow_html=True
    )

    # 1) Selección de referencia de semana
    fecha_ref = st.date_input(
        "Selecciona fecha de la semana",
        dt.date.today(),
        format="DD/MM/YYYY",
        help="Elige un día dentro de la semana que quieres visualizar"
    )

    # 2) Cálculo de Lunes–Viernes
    lunes  = fecha_ref - dt.timedelta(days=fecha_ref.weekday())
    fechas = [lunes + dt.timedelta(days=i) for i in range(5)]
    nombres = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes"]
    cols = [f"{nombres[i]} {fechas[i].strftime(DATE_FMT)}" for i in range(5)]

    # 3) Etiquetas de bloque (solo "Bloque 1", "Bloque 2", …)
    bloque_labels = [f"Bloque {i+1}" for i in range(len(BLOQUES))]

    # 4) Helper para mostrar “Nombre Apellido”
    def short_name(full: str) -> str:
        parts = full.split()
        if len(parts) >= 3:
            return f"{parts[0]} {parts[-2]}"
        elif len(parts) == 2:
            return f"{parts[0]} {parts[1]}"
        return parts[0]

    # 5) DataFrame vacío con índice = bloques, columnas = días
    tabla = pd.DataFrame("", index=bloque_labels, columns=cols)

    # 6) Rellenar cada celda con las reservas, ordenadas y con nombre corto
    for bi, (hi, hf) in enumerate(BLOQUES):
        for di, dia in enumerate(fechas):
            sel = df[
                (df["Fecha"].apply(parse_date) == dia) &
                (df["Hora inicio"].apply(as_time) == hi) &
                (df["Hora fin"].apply(as_time) == hf)
            ]
            if not sel.empty:
                sel_sorted = sel.sort_values(
                    by=["Hora inicio", "Profesor"],
                    key=lambda col: col.apply(as_time) if col.name == "Hora inicio" else col
                )
                lines = sel_sorted.apply(
                    lambda r: (
                        f"{short_name(r['Profesor'])} | "
                        f"{r['Curso']} | {r['Recurso']}"
                    ),
                    axis=1
                ).tolist()
                # Separar cada reserva con una línea en blanco extra
                tabla.iat[bi, di] = "\n\n".join(lines)

    # 7) Estilo con colores y pre-wrap
    import hashlib
    def get_color(text: str) -> str:
        h = hashlib.md5(text.encode("utf-8")).hexdigest()
        return f"#{h[:6]}"

    styled = (
        tabla.style
             .set_properties(**{"white-space": "pre-wrap"})
             .applymap(
                 lambda v: (
                     f"background-color: {get_color(v)}; color:#f4f4f4 !important;"
                     if v else ""
                 )
             )
             .set_table_styles([
                 {"selector": "th", "props": [("min-width", "150px")]},
                 {"selector": "td", "props": [("min-width", "150px")]}
             ])
             .set_table_attributes('style="table-layout:auto; width:100%;"')
    )

    st.markdown(styled.to_html(), unsafe_allow_html=True)

# ------------------------------------------------------------------
# 11) Sección 🔧 Mantenimiento (solo admin)
# ------------------------------------------------------------------
elif page == '🔧 Mantenimiento' and role == 'admin':
    st.markdown(
        "<h2 style='color:#000;font-size:1.5rem;text-align:center'>🔧 Gestión de Mantenimiento</h2>",
        unsafe_allow_html=True
    )

    # — Recalcular listas —
    PROFESORES, CURSOS, RECURSOS = recalc_lists(df)

    def_tab = "Mantenimientos"
    sheets   = pd.ExcelFile(EXCEL).sheet_names

    # — Cargo la hoja de mantenimiento o creo un DataFrame vacío —
    if def_tab in sheets:
        mant_df = pd.read_excel(EXCEL, sheet_name=def_tab).fillna("")
    else:
        mant_df = pd.DataFrame(
            columns=["Recurso", "FechaInicio", "HoraInicio", "FechaFin", "HoraFin"]
        )

    # --- Agregar nuevo mantenimiento ---
    st.subheader("Agregar nuevo mantenimiento")
    rsrc_maint = st.selectbox("Recurso", RECURSOS, key="maint_res")
    d1, d2     = st.columns(2)
    with d1:
        start_date = st.date_input("Fecha inicio", dt.date.today(), key="mant_start_date")
    with d2:
        end_date   = st.date_input("Fecha fin",   dt.date.today(), key="mant_end_date")
    t1, t2     = st.columns(2)
    with t1:
        start_time = st.time_input("Hora inicio", BLOQUES[0][0], key="mant_start_time")
    with t2:
        end_time   = st.time_input("Hora fin",    BLOQUES[0][1], key="mant_end_time")

    if st.button("💾 Guardar mantenimiento", key="save_maint", use_container_width=True):
        new_row = {
            "Recurso":     rsrc_maint,
            "FechaInicio": start_date,
            "HoraInicio":  start_time,
            "FechaFin":    end_date,
            "HoraFin":     end_time,
        }
        mant_df = pd.concat([mant_df, pd.DataFrame([new_row])], ignore_index=True)
        with pd.ExcelWriter(EXCEL, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            mant_df.to_excel(writer, sheet_name=def_tab, index=False)
        st.success("✅ Mantenimiento registrado correctamente.")

    # --- Editar o eliminar mantenimientos ---
    st.subheader("Editar o eliminar mantenimientos")
    mant_editor = mant_df.copy()
    if "FechaInicio" in mant_editor.columns:
        mant_editor["FechaInicio"] = mant_editor["FechaInicio"].apply(parse_date)
    if "FechaFin" in mant_editor.columns:
        mant_editor["FechaFin"]    = mant_editor["FechaFin"].apply(parse_date)
    if "HoraInicio" in mant_editor.columns:
        mant_editor["HoraInicio"]  = mant_editor["HoraInicio"].apply(as_time)
    if "HoraFin" in mant_editor.columns:
        mant_editor["HoraFin"]     = mant_editor["HoraFin"].apply(as_time)

    cfg_maint = {
        "Recurso":      st.column_config.SelectboxColumn("Recurso", options=RECURSOS),
        "FechaInicio":  st.column_config.DateColumn("Fecha inicio", format="DD/MM/YYYY"),
        "HoraInicio":   st.column_config.TimeColumn("Hora inicio", format="HH:mm"),
        "FechaFin":     st.column_config.DateColumn("Fecha fin", format="DD/MM/YYYY"),
        "HoraFin":      st.column_config.TimeColumn("Hora fin", format="HH:mm"),
    }
    edited_maint = st.data_editor(
        mant_editor,
        hide_index=False,
        use_container_width=True,
        column_config=cfg_maint,
    )

    c3, c4 = st.columns(2)
    with c3:
        if st.button("💾 Guardar cambios mantenimiento", key="save_maint_edits", use_container_width=True):
            with pd.ExcelWriter(EXCEL, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                edited_maint.to_excel(writer, sheet_name=def_tab, index=False)
            st.success("✅ Cambios en mantenimiento guardados.")
    with c4:
        to_drop_maint = st.multiselect(
            "Seleccionar mantenimientos a eliminar",
            options=edited_maint.index,
            key="drop_maint",
        )
        if st.button("🗑️ Eliminar mantenimiento", key="delete_maint", use_container_width=True) and to_drop_maint:
            kept = edited_maint.drop(index=to_drop_maint).reset_index(drop=True)
            with pd.ExcelWriter(EXCEL, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                kept.to_excel(writer, sheet_name=def_tab, index=False)
            st.success("✅ Mantenimientos eliminados.")